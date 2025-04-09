import time
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from util_css_selectors import initialize_driver

# Load the Excel workbook and select the active sheet
workbook = load_workbook("Jumbo_product_codes.xlsx")
sheet = workbook.active

print(f'Max Rows Excel < {sheet.max_row} >')

# Iterate through each row in the first column with better failure proof. Precise check.
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):# metoda de iterare mai avansata, cu argumente keyword

    for cell in row:
        if (site_var := cell.value) is None: # Walrus operator. Assign and check at the same time.
            continue

        page_url = f'https://www.e-jumbo.ro/ro/cautare-produs/?search-for={site_var}'
        page = requests.get(page_url)
        soup = BeautifulSoup(page.text, 'html.parser')

        print(f'Accessing page: {page_url}')

        try:
            status = "Status not found"
            if in_magazin_element := soup.find('div', class_="tag tag-9 item"):
                status = in_magazin_element.get_text(strip=True)
            elif doar_online_element := soup.find('div', class_="tag tag-7 item"):
                status = doar_online_element.get_text(strip=True)
            elif online_element := soup.find('div', class_="buy btn js-add-to-cart add-to-cart"):
                status = online_element.get_text(strip=True)
                if status == 'ADAUGA IN COS':
                    status = 'Disponibil'  # Custom status if add to cart is available
            elif no_element := soup.find('div', class_="no-results"):
                status = no_element.get_text(strip=True)

            print(f"Status: {status}")

        except AttributeError as e:
            print(f"An error occurred: {e}")

        # Write the status to the Excel file in the next column
        status_cell = sheet.cell(row=cell.row, column=2)
        status_cell.value = status

        if status == 'Disponibil Numai in Magazine':
            address_data = initialize_driver(page_url)
            i=3
            for elem in address_data:
                address_cell = sheet.cell(row=cell.row, column=i)
                address_cell.value = elem['address']
                i+=1
                stock_cell = sheet.cell(row=cell.row, column=i)
                stock_cell.value = elem['stock']
                i+=1


        # Save the workbook after each update
        workbook.save('Jumbo_product_codes.xlsx')

        # Wait for a bit before processing the next row
        time.sleep(5)

# Keep the browser open until user decides to close it
input("Press Enter to close the browser and exit...")

